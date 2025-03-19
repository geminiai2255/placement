from django.http import FileResponse, HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth import get_user_model
from django.contrib.auth import login,authenticate
from django.core.files.storage import FileSystemStorage
from django.contrib import messages
from django.urls import reverse
from django.views import View
from openpyxl import Workbook
from django.shortcuts import render
from .models import Job
from django.db.models import Q
from .models import JobApplication, Job
from django.http import JsonResponse
from .models import Course, tbl_department
import os
from django.shortcuts import redirect
from openpyxl.worksheet.hyperlink import Hyperlink
from .models import tbl_department
from .models import*
from datetime import datetime, timezone
from .models import SessionApplication, TrainingSession
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import tbl_tutor
from django.shortcuts import render, HttpResponse
from django.db.models import Q
from .models import Job, tbl_student, tbl_tutor, tbl_department


def index(request):
    return render(request, 'index.html')

def login(request):
    full_time_jobs = Job.objects.filter(job_type='full-time')
    part_time_jobs = Job.objects.filter(job_type='part-time')
    featured_jobs = Job.objects.filter(Q(job_type='full-time') | Q(job_type='part-time')).order_by('-date')[:5]

    if request.method == "POST":
        role = request.POST.get('role')  
        pswd = request.POST.get('password')

        # Handle student login
        if role == "student":
            student_id = request.POST.get('student_id')
            student_email = request.POST.get('student_email')

            student_var = tbl_student.objects.filter(student_email=student_email, student_id=student_id)
            if student_var.exists():
                student = student_var.first()
                request.session['id'] = student.id
                return render(request, 'user/user_index.html', {
                    'full_time_jobs': full_time_jobs,
                    'part_time_jobs': part_time_jobs,
                    'featured_jobs': featured_jobs
                })  
            else:
                return HttpResponse("<script>alert('Invalid student credentials!'); window.location='/';</script>")

        # Handle faculty login
        elif role == "faculty":
            faculty_email = request.POST.get('email')

            tutor_var = tbl_tutor.objects.filter(email=faculty_email, password=pswd)
            if tutor_var.exists():
                tutor_instance = tutor_var.first()
                if tutor_instance.status == 'approved':
                    request.session['id'] = tutor_instance.id
                    return render(request, 'tutor/tutor_index.html')
                elif tutor_instance.status == 'pending':
                    return render(request, 'index.html')  

            return HttpResponse("<script>alert('Invalid faculty credentials!'); window.location='/';</script>")

        # Handle admin login
        elif role == "admin":
            admin_email = request.POST.get('email')

            if admin_email == "admin@gmail.com" and pswd == "123":
                request.session['id'] = "admin"  # Use 'id' for consistency
                departments = tbl_department.objects.all()
                return render(request, 'admin/admin_index.html', {'departments': departments})
            else:
                return HttpResponse("<script>alert('Invalid admin credentials!'); window.location='/';</script>")

        else:
            return HttpResponse("<script>alert('Please select a valid role!'); window.location='/';</script>")

    return render(request, "login.html")


def logout(request):
    if request.session.has_key('id'):
        del request.session['id']
        logout(request)
    return render(request,'index.html')

#admin

def add_department(request):
    if request.method == "POST":
        name = request.POST.get('name')
        if tbl_department.objects.filter(name=name).exists():
            messages.error(request, "A department with this name already exists.")
            return redirect('add_department')

        try:
            department = tbl_department.objects.create(name=name)
            messages.success(request, "Department added successfully!")
            return render(request, 'admin/add_department.html', {'success': True})
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect('add_department')
    return render(request, 'admin/add_department.html')


def list_department(request):
    if 'id' not in request.session:
        return redirect('/login/') 
    departments = tbl_department.objects.all()
    return render(request, 'admin/list_department.html', {'data': departments})

def admin_index(request):
    departments = tbl_department.objects.all() 
    return render(request, 'admin/admin_index.html', {'departments': departments})


from django.shortcuts import render
from .models import tbl_tutor
def tutor_index(request):
    tutor_id = request.session.get('id')
    if tutor_id:
        try:
            tutor = tbl_tutor.objects.get(id=tutor_id)
            full_time_jobs = Job.objects.filter(job_type='full-time', department=tutor.department)
            part_time_jobs = Job.objects.filter(job_type='part-time', department=tutor.department)
            featured_jobs = Job.objects.filter(
                Q(job_type='full-time', department=tutor.department) |
                Q(job_type='part-time', department=tutor.department)
            ).order_by('-date')[:5]

            unseen_notifications_count = TutorNotification.objects.filter(tutor=tutor, is_read=False).count()

            context = {
                'full_time_jobs': full_time_jobs,
                'part_time_jobs': part_time_jobs,
                'featured_jobs': featured_jobs,
                'notifications_count': unseen_notifications_count,
            }
        except tbl_tutor.DoesNotExist:
            request.session.flush()
            context = {'error': 'Tutor data not found.'}
    else:
        context = {'error': 'Please log in to access this page.'}

    return render(request, 'tutor/tutor_index.html', context)

from django.shortcuts import render
from .models import Job

def user_index(request):
    student_id = request.session.get('id')
    student = tbl_student.objects.get(id=student_id)
    full_time_jobs = Job.objects.filter(job_type='full-time',department_id=student.department)
    part_time_jobs = Job.objects.filter(job_type='part-time',department_id=student.department)
    featured_jobs = Job.objects.filter(Q(job_type='full-time',department_id=student.department) | Q(job_type='part-time',department_id=student.department)).order_by('-date')[:5]
   

    # Pass the jobs to the template
    return render(request, 'user/user_index.html', {
        'full_time_jobs': full_time_jobs,
        'part_time_jobs': part_time_jobs,
        'featured_jobs': featured_jobs
    })

def job_list(request):
    return render(request, 'user/job_list.html')

def job_detail(request):
    return render(request, 'user/job_detail.html')

def user_profile(request):
    if 'id' not in request.session:
        return redirect('/login/')

    user_id = request.session['id']  
    try:
        user = tbl_student.objects.get(id=user_id)  
    except tbl_student.DoesNotExist:
        return redirect('/login/')  
    departments = tbl_department.objects.all()

    if request.method == 'POST':
        user.name = request.POST.get('name', user.name)
        user.email = request.POST.get('email', user.email)
        user.gender = request.POST.get('gender', user.gender)
        user.mobile_number = request.POST.get('mobile_number', user.mobile_number)
        selected_department_name = request.POST.get('department')
        if selected_department_name:
            selected_department = tbl_department.objects.get(name=selected_department_name)
            user.department = selected_department
        
        user.save()
        return redirect('/user_profile/')  
    return render(request, 'user/user_profile.html', {'data': user, 'departments': departments})

def user_profile(request):
    if 'id' not in request.session:
        return redirect('/login/')

    user_id = request.session['id']
    try:
        user = tbl_student.objects.get(id=user_id)
    except tbl_student.DoesNotExist:
        return redirect('/login/')

    departments = tbl_department.objects.all()

    if request.method == 'POST':
        user.name = request.POST.get('name', user.name)
        user.student_email = request.POST.get('email', user.student_email)
        user.gender = request.POST.get('gender', user.gender)
        user.student_phone = request.POST.get('mobile_number', user.student_phone)
        user.address = request.POST.get('address', user.address)
        user.date_of_birth = request.POST.get('date_of_birth', user.date_of_birth)

        selected_department_name = request.POST.get('department')
        if selected_department_name:
            try:
                selected_department = tbl_department.objects.get(name=selected_department_name)
                user.department = selected_department
            except tbl_department.DoesNotExist:
                pass

        # Save updated user profile
        user.save()

        return redirect('/user_profile/')  
    return render(request, 'user/user_profile.html', {'data': user, 'departments': departments})




def admin_view_tutor(request):
    data = tbl_tutor.objects.filter(status='pending')
    return render(request, 'admin/admin_view_tutor.html', {'data': data})


def admin_approved_tutor(request):
    id=request.GET['id']
    tbl_tutor.objects.all().filter(id=id).update(status='approved') 
    return HttpResponseRedirect('/admin_view_tutor/')


def admin_view_approved_tutor(request):
    data=tbl_tutor.objects.all().filter(status='approved')
    return render(request,'admin/admin_approved_tutor.html',{'data':data})


def admin_rejected_tutor(request):
    id=request.GET['id']
    tbl_tutor.objects.all().filter(id=id).update(status='rejected') 
    return render(request,'admin/admin_rejected_tutor.html')


def admin_view_rejected_tutor(request):
    data=tbl_tutor.objects.all().filter(status='Rejected')
    return render(request,'admin/admin_rejected_tutor.html',{'data':data})

def session_applylist(request, session_id):
    # Get the session object by ID
    session = TrainingSession.objects.get(id=session_id)
    session_applications = SessionApplication.objects.filter(session=session)
    applied_students = [application.user for application in session_applications]

    # Render the page with the list of applied students
    return render(request, 'tutor/session_applylist.html', {
        'data': applied_students,
        'session': session
    })
    

def list_student(request):
    tutor_id = request.session.get('id')
    if not tutor_id:
        messages.error(request, "Session expired or tutor not logged in.")
        return redirect('login')

    tutor = tbl_tutor.objects.get(id=tutor_id)
    
    # Check if course_id filter is the issue
    data = tbl_student.objects.filter(course_id=tutor.course_id)

    # Debugging print to ensure data is loaded
    print(f"Filtered Data: {data}")
    
    if not data.exists():  # Display all data for testing
        data = tbl_student.objects.all()  
        print(f"All Data: {data}")

    return render(request, 'tutor/list_student.html', {'data': data})




def tutor_jobapplylist(request, job_id):
    job = get_object_or_404(Job, id=job_id)
    tutor_id = request.session.get('id')
    tutor = get_object_or_404(tbl_tutor, id=tutor_id)

    # Get the tutor's course
    course = tutor.course  

    # Fetch students enrolled in the same course
    students = tbl_student.objects.filter(course_id=course.id)

    # Filter job applications only from these students
    applications = JobApplication.objects.filter(job=job, user__in=students)

    # Pass data to the template
    return render(request, 'tutor/tutor_jobapplylist.html', {
        'job': job,
        'data': applications  # Only applications from students in the tutor's course
    })
    
from django.shortcuts import render, get_object_or_404, redirect
from .models import Job, JobApplication, tbl_student, tbl_tutor



def delete_student(request):
    if 'id' in request.GET:
        student_id = request.GET['id']
        try:
            student = tbl_student.objects.get(id=student_id)
            student.delete()
            return redirect('list_student')  
        except tbl_student.DoesNotExist:
            return render(request, 'tutor/add_student.html', {'error': 'Student not found'})
    else:
        return redirect('list_student')  
 
    
def delete_department(request):
    if 'id' in request.GET:
        department_id = request.GET['id']
        try:
            department = tbl_department.objects.get(id=department_id)
            department.delete()
            return redirect('list_department')  
        except tbl_department.DoesNotExist:
            return render(request, 'admin/department_list.html', {'error': 'Department not found'})
    else:
        return redirect('list_department') 


def view_department(request):
    departments = tbl_department.objects.all()  
    return render(request, 'admin/view_department.html', {'data': departments})
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import Job, tbl_department
from django.http import JsonResponse
import traceback
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from .models import Job, tbl_department
from django.shortcuts import render, get_object_or_404
from .models import tbl_student, tbl_department, Course
def edit_profile(request, student_id):
    student = get_object_or_404(tbl_student, id=student_id)
    departments = tbl_department.objects.all()
    courses = Course.objects.all()

    if request.method == 'POST':
        # Collect required fields
        name = request.POST.get('name')
        date_of_birth = request.POST.get('date_of_birth')
        gender = request.POST.get('gender')
        email = request.POST.get('email')
        roll_no = request.POST.get('roll_no')
        reg_no = request.POST.get('reg_no')
        student_id = request.POST.get('student_id')
        department_id = request.POST.get('department')
        course_id = request.POST.get('course')

        # Validate required fields
        if not all([name, date_of_birth, gender, email, roll_no, reg_no, student_id, department_id, course_id]):
            return render(request, 'user/edit_profile.html', {
                'student': student,
                'departments': departments,
                'courses': courses,
                'error': "All fields are required."
            })

        # Update personal details
        student.name = name
        student.date_of_birth = date_of_birth
        student.gender = gender
        student.student_email = email
        student.roll_no = roll_no
        student.reg_no = reg_no

        # Update department and course
        student.department = tbl_department.objects.get(id=department_id)
        student.course = Course.objects.get(id=course_id)

        # Update optional fields
        student.student_phone = request.POST.get('student_phone', student.student_phone)
        student.parent_phone = request.POST.get('parent_phone', student.parent_phone)
        student.year = request.POST.get('year', student.year)
        student.batch = request.POST.get('batch', student.batch)
        student.supply = request.POST.get('supply', student.supply)
        student.pancard_number = request.POST.get('pancard_number', student.pancard_number)

        # Handle file uploads
        if 'image' in request.FILES:
            student.image = request.FILES['image']
        if 'resume' in request.FILES:
            student.resume = request.FILES['resume']

        # Update password if provided
        password = request.POST.get('password')
        if password:
            student.password = password

        # Save the updated student record
        student.save()

        # Redirect with a success flag
        return render(request, 'user/edit_profile.html', {
            'student': student,
            'departments': departments,
            'courses': courses,
            'success': True
        })

    # Render the form with current data
    return render(request, 'user/edit_profile.html', {
        'student': student,
        'departments': departments,
        'courses': courses
    })

    
def edit_department(request, department_id):
    if 'id' not in request.session:
        return redirect('/login/')  
    try:
        department = tbl_department.objects.get(id=department_id)
    except tbl_department.DoesNotExist:
        messages.error(request, "Department not found")
        return redirect('list_department')  
    if request.method == 'POST':
        department.name = request.POST.get('name')  
        department.save()

        messages.success(request, "Department updated successfully")
        return redirect('list_department')  
    return render(request, 'admin/edit_department.html', {'department': department})


def admin_joblist(request):
    jobs = Job.objects.all()  
    return render(request, 'admin/admin_joblist.html', {'data': jobs})


def admin_jobdetail(request):
    return render(request, 'admin/admin_jobdetail.html')


from django.core.paginator import Paginator
from django.shortcuts import render
from .models import Job  # Make sure to import your Job model

def user_job_list(request):
    student_id = request.session.get('id')
    student = tbl_student.objects.get(id=student_id)
    full_time_jobs = Job.objects.filter(job_type='full-time',department_id=student.department)
    part_time_jobs = Job.objects.filter(job_type='part-time',department_id=student.department)
    featured_jobs = Job.objects.filter(Q(job_type='full-time',department_id=student.department) | Q(job_type='part-time',department_id=student.department)).order_by('-date')[:5]
    return render(request, 'user/user_job_list.html', {
        'full_time_jobs': full_time_jobs,
        'part_time_jobs': part_time_jobs,
        'featured_jobs': featured_jobs
    })



def search_jobs(request):
    search_query = request.GET.get('search', '')
    jobs = Job.objects.filter(
        Q(description__icontains=search_query) |
        Q(job_type__icontains=search_query) |
        Q(place__icontains=search_query)
    )
    job_list = []
    for job in jobs:
        job_list.append({
            'id': job.id,
            'title': job.title,
            'place': job.place,
            'job_type': job.job_type,
            'salary': job.salary,
            'date': job.date,
            'logo': job.logo.url if job.logo else '',  
            'department': job.department.name if job.department else 'N/A',  
        })

    return JsonResponse({'jobs': job_list})


def user_job_detail(request, job_id):
    job = get_object_or_404(Job, id=job_id)
    return render(request, 'user/user_job_detail.html', {'job': job})

def tutor_job_list(request):
    tutor_id = request.session.get('id')
    tutor = tbl_tutor.objects.get(id=tutor_id)
    full_time_jobs = Job.objects.filter(job_type='full-time',department=tutor.department)
    part_time_jobs = Job.objects.filter(job_type='part-time',department=tutor.department)
    featured_jobs = Job.objects.filter(Q(job_type='full-time',department=tutor.department) | Q(job_type='part-time',department=tutor.department)).order_by('-date')[:5]
    return render(request, 'tutor/tutor_job_list.html', {
        'full_time_jobs': full_time_jobs,
        'part_time_jobs': part_time_jobs,
        'featured_jobs': featured_jobs
    })
    
def tutor_job_detail(request, id):
    job = get_object_or_404(Job, id=id)  
    return render(request, 'tutor/tutor_job_detail.html', {'job': job})

from django.shortcuts import get_object_or_404, render
from django.http import JsonResponse
def apply_for_job(request, job_id):
    if request.method == 'POST':
        # Get the student object based on the logged-in user
        student = request.user.tbl_student

        # Get the job object based on job_id
        job = get_object_or_404(Job, id=job_id)

        # Check if the student has already applied for this job
        existing_application = JobApplication.objects.filter(user=student, job=job).exists()
        if existing_application:
            return JsonResponse({'success': False, 'message': 'You have already applied for this job.'})

        # Create the application
        job_application = JobApplication.objects.create(user=student, job=job)

        # Create a notification for the tutor
        tutor = job.course.tutor  # Assuming each job is linked to a course and the course has a tutor

        # Corrected variable name and added notification save
        notification = TutorNotification(
            course=job.course,
            student=student,
            job=job,
            notification_type='job_application',
            message=f'{student.name} has applied for the job: {job.title}',
            is_read=False
        )
        notification.save()  # Save the notification

        return JsonResponse({'success': True, 'message': 'Successfully applied for the job!'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})
from django.contrib import messages
from django.shortcuts import redirect, render
from datetime import datetime

def admin_training_sessions(request):
    if request.method == 'POST':
        session_name = request.POST.get('session_name')
        description = request.POST.get('description')
        trainer_name = request.POST.get('trainer_name')
        venue = request.POST.get('venue')
        time = request.POST.get('time')
        training_dates_str = request.POST.get('training_dates')

        if not session_name or not description or not trainer_name or not venue or not time or not training_dates_str:
            messages.error(request, "All fields are required!")
            return render(request, 'admin/admin_training_sessions.html')

        try:
            # Parse and validate dates
            training_dates_list = [date.strip() for date in training_dates_str.split(',')]
            today = datetime.today().date()
            invalid_dates = [
                date for date in training_dates_list if datetime.strptime(date, '%Y-%m-%d').date() < today
            ]

            if invalid_dates:
                return render(
                    request,
                    'admin/admin_training_sessions.html',
                    {'alert_message': f"Invalid dates: {', '.join(invalid_dates)}. Dates must be today or in the future."}
                )
        except ValueError:
            return render(
                request,
                'admin/admin_training_sessions.html',
                {'alert_message': "Error in training dates format. Use YYYY-MM-DD format."}
            )

        # Save valid data
        training_session = TrainingSession(
            session_name=session_name,
            description=description,
            trainer_name=trainer_name,
            venue=venue,
            time=time,
            training_dates=training_dates_list
        )
        training_session.save()
        
        # Add success message
        return render(
            request,
            'admin/admin_training_sessions.html',
            {'success_message': "Session added successfully!"}
        )

    return render(request, 'admin/admin_training_sessions.html')

def admin_sessionlist(request):
    training_sessions = TrainingSession.objects.all()
    return render(request, 'admin/admin_sessionlist.html', {'training_sessions': training_sessions})



def delete_TrainingSession(request, session_id):
    try:
        training_session = TrainingSession.objects.get(id=session_id)
        training_session.delete()
        return redirect('admin_sessionlist') 
    except TrainingSession.DoesNotExist:
        return render(request, 'admin/admin_training_sessions.html', {'error': 'Training session not found'})


def admin_editsession(request, session_id):
    if 'id' not in request.session:
        return redirect('/login/')  
    training_session = get_object_or_404(TrainingSession, id=session_id)

    if request.method == 'POST':
        time_str = request.POST.get('time')
        try:
            time_obj = datetime.strptime(time_str, '%I:%M %p').time()
        except ValueError:
            messages.error(request, "Invalid time format. Please use HH:MM AM/PM.")
            return render(request, 'admin/admin_editsession.html', {'TrainingSession': training_session})
        training_session.session_name = request.POST.get('session_name')
        training_session.description = request.POST.get('description')
        training_session.trainer_name = request.POST.get('trainer_name')
        training_session.venue = request.POST.get('venue')
        training_session.time = time_obj
        training_session.save()
        messages.success(request, "Session updated successfully")
        return redirect('admin_sessionlist')  

    return render(request, 'admin/admin_editsession.html', {'TrainingSession': training_session})
def admin_applylist(request, job_id):
    job_applications = JobApplication.objects.filter(job_id=job_id).select_related('user')
    return render(request, 'admin/admin_applylist.html', {'data': job_applications})
def delete_job(request, job_id):
    try:
        job = Job.objects.get(id=job_id)
        job.delete()
        return redirect('admin_joblist')  # Change to the correct job list view name
    except Job.DoesNotExist:
        return render(request, 'admin/admin_joblist.html', {'error': 'Job not found'})
def tutor_applylist(request):
    if 'id' not in request.session:
        return redirect('/tutor_applylist.html/')  
    data = JobApplication.objects.all() 
    return render(request, 'tutor/tutor_applylist.html', {'data': data})


def user_about(request):
    return render(request, 'user/user_about.html')

def tutor_about(request):
    return render(request, 'tutor/tutor_about.html')

def contact_view(request):
    return render(request, 'user/contact.html')


def add_course(request, department_id=None):
    if request.method == "POST":
        name = request.POST.get('name')
        
        if Course.objects.filter(name=name, department_id=department_id).exists():
            messages.error(request, "A Course with this name already exists in the department.")
            return redirect('add_course', department_id=department_id)

        try:
            department = get_object_or_404(tbl_department, id=department_id)
            course = Course.objects.create(
                name=name,
                department=department
            )
            course.save()

            messages.success(request, "Course added successfully!")
            return redirect('courses_listing', department_id=department_id)  

        except tbl_department.DoesNotExist:
            messages.error(request, "The selected department does not exist.")
            return redirect('list_department')  
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect('add_course', department_id=department_id)

    if department_id:
        try:
            department = tbl_department.objects.get(id=department_id)
            return render(request, 'admin/add_course.html', {'department': department})
        except tbl_department.DoesNotExist:
            messages.error(request, "The department does not exist.")
            return redirect('list_department')
    else:
        return redirect('list_department')  



from django.shortcuts import render

def courses_listing(request, department_id):
    try:
        department = tbl_department.objects.get(id=department_id)
        courses = Course.objects.filter(department=department)

        # Dictionary to store batches for each tutor inside each course
        course_batches = {}

        for course in courses:
            tutors_batches = {}
            for tutor in course.tutors.all():
                batch_list = [batch.strip() for batch in tutor.batch.split(',')]
                tutors_batches[tutor.id] = batch_list  # Store batches for each tutor

            course_batches[course.id] = tutors_batches  # Store tutor batches per course

    except tbl_department.DoesNotExist:
        return render(request, 'admin/courseslisting.html', {'error': 'Department not found'})

    return render(request, 'admin/courseslisting.html', {
        'department': department,
        'courses': courses,
        'course_batches': course_batches,  # Pass batches grouped by course and tutor
    })

def delete_course(request, id):
    try:
        course = Course.objects.get(id=id)
        department_id = course.department.id  
        course.delete()
        return redirect('list_course', department_id=department_id)  
    except Course.DoesNotExist:
        return render(request, 'admin/list_course.html', {'error': 'Course not found'})



def tutor_sessionlist(request):
    training_sessions = TrainingSession.objects.all()
    return render(request, 'tutor/tutor_sessionlist.html', {'training_sessions': training_sessions})


def user_sessionlist(request):
    training_sessions = TrainingSession.objects.all()
    return render(request, 'user/user_sessionlist.html', {'training_sessions': training_sessions})
from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse
from .models import TrainingSession, SessionApplication

def apply_for_session(request):
    if request.method == 'POST':
        user_id = request.session.get('id')
        if not user_id:
            return JsonResponse({'success': False, 'message': 'User not logged in.'})

        student = get_object_or_404(tbl_student, id=user_id)
        session_id = request.POST.get('session_id')
        session = get_object_or_404(TrainingSession, id=session_id)

        # Check if already applied
        existing_application = SessionApplication.objects.filter(user=student, session=session).first()
        if existing_application:
            return JsonResponse({'success': False, 'message': 'You have already applied for this session.'})

        # Create the application with current datetime
        SessionApplication.objects.create(user=student, session=session, applied_on=datetime.now())

        return JsonResponse({'success': True, 'message': 'Successfully applied for the session!'})


def user_applied_sessions(request):
    # Get the user_id from the session
    user_id = request.session.get('id')

    if not user_id:
        return render(request, 'user/user_applied_sessions.html', {'applied_sessions': []})

    # Get the student object
    student = get_object_or_404(tbl_student, id=user_id)

    # Fetch sessions the student has applied for
    applied_sessions = SessionApplication.objects.filter(user=student).select_related('session')

    return render(request, 'user/user_applied_sessions.html', {'applied_sessions': applied_sessions})

def apply_for_job(request, job_id):
    # Get the user_id from the session
    user_id = request.session.get('id')

    # Check if user_id is available in the session
    if not user_id:
        return JsonResponse({'success': False, 'message': 'You must be logged in to apply for a job.'})

    try:
        # Retrieve the student object using the session user_id
        student = get_object_or_404(tbl_student, id=user_id)
    except tbl_student.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Student information not found for this user.'})

    # Get the job object based on job_id
    job = get_object_or_404(Job, id=job_id)

    # Check if the student has already applied for the job
    existing_application = JobApplication.objects.filter(user=student, job=job).exists()
    if existing_application:
        return JsonResponse({'success': False, 'message': 'You have already applied for this job.'})

    # Create a new application
    JobApplication.objects.create(user=student, job=job)
    return JsonResponse({'success': True, 'message': 'Successfully applied for the job!'})


from django.shortcuts import render, redirect
from .models import JobApplication
def user_applied_jobs(request):
    # Get the user_id from the session
    user_id = request.session.get('id')

    # If there's no user_id in the session, return an empty list for applied jobs
    if not user_id:
        applied_jobs = []  # No job applications for an anonymous user
    else:
        try:
            # Retrieve the student object using the session user_id
            student = get_object_or_404(tbl_student, id=user_id)
            # Fetch all job applications for the student
            applied_jobs = JobApplication.objects.filter(user=student)
        except tbl_student.DoesNotExist:
            applied_jobs = []  # If student doesn't exist, return empty list

    # Return the jobs and the applications to the template
    return render(request, 'user/user_applied_job.html', {'applied_jobs': applied_jobs})
def admin_applysessionlist(request, session_id):
    # Fetch applications for the selected session
    applications = SessionApplication.objects.select_related('user', 'session').filter(session_id=session_id)
    
    # Fetch the session name for context
    session = TrainingSession.objects.get(id=session_id)
    
    return render(request, 'admin/admin_applysessionlist.html', {
        'data': applications,
        'session_name': session.session_name
    })
    
def tutor_profile(request):
    if 'id' not in request.session:
        return redirect('/login/')  
    tutor_id = request.session['id']  
    try:
        tutor = tbl_tutor.objects.get(id=tutor_id)  
    except tbl_tutor.DoesNotExist:
        return redirect('/login/')  

    if request.method == 'POST':
        tutor.name = request.POST.get('name')
        tutor.email = request.POST.get('email')  
        tutor.password = request.POST.get('password')
        tutor.save()  
        return redirect('/tutor_profile/')  
    return render(request, 'tutor/tutor_profile.html', {'data': tutor})  

def tutor_editprofile(request):
    if 'id' not in request.session:
        return redirect('/login/')  # Redirect to login if the user is not authenticated

    tutor_id = request.session['id']
    tutor = get_object_or_404(tbl_tutor, id=tutor_id)  # Retrieve the Tutor object

    # Fetch all departments and courses from the database
    departments = tbl_department.objects.all()
    courses = Course.objects.all()

    if request.method == 'POST':
        # Update fields from the POST request
        tutor.name = request.POST.get('name', tutor.name)
        tutor.email = request.POST.get('email', tutor.email)

        # Update password if provided
        password = request.POST.get('password')
        if password:
            tutor.password = password  # Directly assign the password (not hashed)

        tutor.gender = request.POST.get('gender', tutor.gender)
        tutor.mobile_number = request.POST.get('mobile_number', tutor.mobile_number)

        # Fetch the department object based on the selected department ID
        department_id = request.POST.get('department')
        if department_id:
            tutor.department = tbl_department.objects.get(id=department_id)

        # Fetch the course object based on the selected course ID
        course_id = request.POST.get('course')
        if course_id:
            tutor.course = Course.objects.get(id=course_id)

        # Handle multiple batch periods as comma-separated values
        tutor.batch = request.POST.get('batch', tutor.batch)

        # Handle file upload for profile image
        if request.FILES.get('image'):
            tutor.image = request.FILES['image']

        # Save the updated Tutor object
        tutor.save()

        # Display a success message and redirect to the profile page
        messages.success(request, 'Profile updated successfully!')
        return redirect('tutor_profile')  # Replace with the correct URL name for the profile page

    # Render the edit profile page with the current tutor data, departments, and courses
    context = {
        'data': tutor,
        'departments': departments,
        'courses': courses,
    }
    return render(request, 'tutor/tutor_editprofile.html', context)


def download_excel(request):
    # Extract `session_id` from the request
    session_id = request.GET.get('session_id')
    if not session_id:
        return HttpResponse("Error: 'session_id' parameter is missing.", status=400)
    
    try:
        # Fetch the TrainingSession object
        session = TrainingSession.objects.get(id=session_id)
    except TrainingSession.DoesNotExist:
        return HttpResponse(f"Error: No session found with ID {session_id}.", status=404)
    
    # Fetch all applications for the given session
    applications = SessionApplication.objects.filter(session=session)
    
    # Initialize Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"Applications for {session.session_name}"
    
    # Write the header row
    ws.append(['Sl. No.', 'Student Name', 'Email', 'Contact', 'Program', 'Year', 'Session Name'])
    
    # Write application data rows
    for idx, application in enumerate(applications, start=1):
        user = application.user  # Access the user (student) linked to the application
        ws.append([
            idx,
            user.name,  # Student name
            user.student_email,  # Student email
            user.student_phone,  # Student phone
            user.course.name if user.course else 'N/A',  # Course name (if exists)
            user.year if user.year else 'N/A',  # Year (if exists)
            session.session_name  # Training session name
        ])
    
    # Prepare the HTTP response with Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="applications_{session.session_name.replace(" ", "_")}.xlsx"'
    wb.save(response)

    return response

def download_resume(request, student_id):
    student = get_object_or_404(tbl_student, id=student_id)

    if not student.resume:
        return HttpResponse("Error: Resume not available.", status=404)

    # Open the file and send it as response
    resume_path = student.resume.path  # Path to the resume file
    response = FileResponse(open(resume_path, 'rb'), as_attachment=True, filename=student.resume.name)

    return response
def admin_studentslist(request):
      departments = tbl_department.objects.all() 
      return render(request, 'admin/admin_studentslist.html', {'departments': departments})

import openpyxl
from django.http import HttpResponse
from openpyxl.drawing.image import Image
from io import BytesIO
from PIL import Image as PILImage
import fitz  

def download_applications_excel(request, job_id):
    try:
        # Fetch the job and its associated applications
        job = Job.objects.get(id=job_id)
    except Job.DoesNotExist:
        return HttpResponse(f"Error: No job found with ID {job_id}.", status=404)

    applications = JobApplication.objects.filter(job=job).select_related('user')

    # Create a workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{job.title} Applications"

    # Add header row
    headers = [
        "Job Name", "Student Name", "Course", "Date of Birth", "Department", "Resume", 
        "Contact", "Pancard Number", "Email", "Gender", "Batch"
    ]
    ws.append(headers)

    # Populate data rows
    for app in applications:
        student = app.user
        row = [
            job.title,
            student.name,
            student.course.name if student.course else "N/A",
            student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else "N/A",
            student.department.name if student.department else "N/A",
            "",  # Placeholder for resume image
            student.student_phone,
            student.pancard_number if student.pancard_number else "N/A",
            student.student_email,
            student.gender,
            student.batch if student.batch else "N/A",
        ]
        ws.append(row)

        # Add the resume as an image (if PDF, convert to image)
        if student.resume:
            try:
                resume_path = student.resume.path
                if resume_path.endswith('.pdf'):
                    # Convert first page of PDF to an image
                    doc = fitz.open(resume_path)
                    pix = doc[0].get_pixmap()
                    image_bytes = BytesIO(pix.tobytes("png"))
                    img = Image(image_bytes)
                else:
                    # If resume is already an image
                    img = Image(resume_path)

                img.width, img.height = 100, 100  # Resize image
                cell_address = f"F{ws.max_row}"  # Assuming column 'F' for the resume
                ws.add_image(img, cell_address)
            except Exception as e:
                print(f"Error processing resume for {student.name}: {e}")
                # Leave the cell empty if there's an issue

    # Adjust column widths dynamically
    for col in ws.columns:
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Prepare the HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{job.title.replace(" ", "_")}_applications.xlsx"'

    # Save the workbook to the response
    wb.save(response)
    return response

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.http import HttpResponse
from .models import tbl_tutor, tbl_department, Course
import logging

# Get an instance of a logger
logger = logging.getLogger(__name__)

def register_tutor(request):
    if request.method == 'GET':
        # Fetch departments and courses to populate dropdowns
        departments = tbl_department.objects.all()
        courses = Course.objects.all()
        
        return render(request, 'register.html', {
            'departments': departments,
            'courses': courses
        })
    
    elif request.method == 'POST':
        # Debug: Log form data
        logger.debug(f"Form Data: {request.POST}")
        logger.debug(f"Uploaded Files: {request.FILES}")

        # Capture form data
        name = request.POST.get('name')
        email = request.POST.get('email')
        gender = request.POST.get('gender')
        mobile_number = request.POST.get('mobile_number')
        password = request.POST.get('password')  # Plain password
        department_id = request.POST.get('department')
        course_id = request.POST.get('course')
        batch = request.POST.get('batch')
        image = request.FILES.get('image')  # Handle file upload if provided

        # Validate and create a new tutor
        try:
            # Check if the email already exists
            if tbl_tutor.objects.filter(email=email).exists():
                messages.error(request, 'This email is already registered. Please use a different email.')
                return redirect('register_tutor')

            # Check if department and course exist
            department = tbl_department.objects.get(id=department_id)
            course = Course.objects.get(id=course_id)

            # Create the tutor instance without password hashing
            new_tutor = tbl_tutor.objects.create(
                name=name,
                email=email,
                gender=gender,
                mobile_number=mobile_number,
                password=password,  # Store plain password
                department=department,
                course=course,
                batch=batch,
                image=image,  # Optional
                status='pending',  # Default value
            )
            new_tutor.save()

            # Manually store tutor information in the session
            request.session['tutor_id'] = new_tutor.id
            request.session['tutor_name'] = new_tutor.name
            request.session['tutor_email'] = new_tutor.email

            # Optionally set session expiration or timeout if needed
            request.session.set_expiry(3600)  # Set session to expire in 1 hour (3600 seconds)

            # Add success message
            messages.success(request, 'Registration successful! You can now log in.')

            # Redirect to the login page or a desired page
            return redirect('login')  # Replace 'login' with the actual URL name of the login page

        except Exception as e:
            # Log the exception and display a generic error message
            logger.error(f"Error occurred: {e}")
            return HttpResponse(f"An error occurred: {e}", status=400)
        
from django.shortcuts import redirect, render
from django.contrib import messages
import pandas as pd  # Install pandas if not already installed

from django.shortcuts import render, redirect
from django.contrib import messages
import pandas as pd
from .models import tbl_student  # Update this with your actual model import


import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import tbl_student # Replace with your actual models
def upload_excel(request):
    if request.method == 'POST' and 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']

        try:
            df = pd.read_excel(excel_file, engine='openpyxl')
            print(df.head())  # Debugging: Check file content

            required_columns = [
                'Student ID', 'Roll No', 'Reg No', 'Student Name', 'Gender',
                'Student Email', 'Student phone', 'Parent phone', 'Address', 'Course'
            ]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ValueError(f"Missing columns: {', '.join(missing_columns)}")

            inserted_count = 0
            updated_count = 0

            for _, row in df.iterrows():
                try:
                    student_id = row['Student ID']

                    # Get course object
                    course = Course.objects.get(name=row['Course'])

                    defaults = {
                        'roll_no': row['Roll No'],
                        'reg_no': row['Reg No'],
                        'name': row['Student Name'],
                        'gender': row['Gender'],
                        'student_email': row['Student Email'],
                        'student_phone': row['Student phone'],
                        'parent_phone': row['Parent phone'],
                        'address': row['Address'],
                        'course': course  # Correcting course assignment
                    }
                    student, created = tbl_student.objects.update_or_create(
                        student_id=student_id,
                        defaults=defaults
                    )
                    if created:
                        print(f"Inserted: {student}")  # Debugging
                        inserted_count += 1
                    else:
                        print(f"Updated: {student}")  # Debugging
                        updated_count += 1

                except Exception as e:
                    print(f"Error processing row: {row}, Error: {e}")
                    continue

            messages.success(
                request,
                f"File uploaded successfully! {inserted_count} new records added and {updated_count} records updated."
            )
            return redirect('list_student')

        except Exception as e:
            messages.error(request, f"An error occurred: {e}")
            print(f"Error: {e}")  # Debugging

        return redirect('list_student')

    return render(request, 'tutor/list_student.html')


from django.shortcuts import render
from .models import tbl_student, Course

def view_final_year_student(request):
    # Get the course_id and batch from query parameters
    course_id = request.GET.get('id')  # Get the course ID from the query params
    batch = request.GET.get('batch')  # Get the batch from the query params

    # Ensure both course_id and batch are provided
    if not course_id or not batch:
        return render(request, 'admin/admin_studentslist.html', {
            'error': 'Both Course ID and Batch are required.',
        })

    try:
        # Fetch the Course object based on course_id
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return render(request, 'admin/admin_studentslist.html', {
            'error': 'Course not found.',
        })

    # Filter students based on course and batch
    students = tbl_student.objects.filter(course=course, batch=batch)

    # Return the rendered template with student data
    return render(request, 'admin/admin_studentslist.html', {
        'students': students,
    })
from django.utils.timezone import now
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.timezone import now
from .models import JobApplication, StudentNotification, Job
from django.shortcuts import render, get_object_or_404, redirect
from django.utils.timezone import now
from .models import JobApplication, StudentNotification
from django.utils.timezone import now
from django.shortcuts import get_object_or_404, redirect

def approve_application(request, application_id):
    job_application = get_object_or_404(JobApplication, id=application_id)
    job_application.status = 'Approved'
    job_application.save()

    student = job_application.user
    job = job_application.job

    # Store student ID in session
    request.session['student_id'] = student.id  

    # Create a notification for the student
    StudentNotification.objects.create(
        student=student,
        job=job,
        message=f"Your application for the job '{job.title}' has been approved.",
        created_at=now()
    )

    return redirect('tutor_jobapplylist', job_id=job.id)

def reject_application(request, application_id):
    job_application = get_object_or_404(JobApplication, id=application_id)
    job_application.status = 'Rejected'
    job_application.save()

    student = job_application.user
    job = job_application.job

    # Store student ID in session
    request.session['student_id'] = student.id  

    # Create a notification for the student
    StudentNotification.objects.create(
        student=student,
        job=job,
        message=f"Your application for the job '{job.title}' has been rejected.",
        created_at=now()
    )

    return redirect('tutor_jobapplylist', job_id=job.id)

def reapprove_application(request, application_id):
    application = get_object_or_404(JobApplication, id=application_id)
    application.status = "Approved"
    application.save()
    messages.success(request, "Application has been reapproved.")
    return redirect('tutor_jobapplylist', job_id=application.job.id)
from django.shortcuts import render
from placementapps.models import StudentNotification

def student_notifications(request):
    student_id = request.session.get('student_id')  # Get student ID from session

    if not student_id:  # If no student ID is found, handle gracefully
        return render(request, 'user/notifications.html', {'notifications': []})

    student = tbl_student.objects.get(id=student_id)  # Fetch student instance

    # **Get jobs where student is "Approved" or "Rejected"**
    applied_jobs = JobApplication.objects.filter(
        user=student_id, 
        status__in=['Approved', 'Rejected']  
    ).values_list('job_id', flat=True)

    
    notifications = StudentNotification.objects.filter(
        student=student, 
        job_id__in=applied_jobs
    ).order_by('-created_at')

    
    notifications.filter(is_read=False).update(is_read=True)

    return render(request, 'user/notifications.html', {'notifications': notifications})
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.shortcuts import render



def admin_addjoblist(request):
    if request.method == 'POST':
        try:
           
            title = request.POST.get('title', '').strip()
            department_id = request.POST.get('department', '').strip()

            if not department_id or not title:
                return JsonResponse({'success': False, 'error': 'Required fields are missing'})

            department = get_object_or_404(tbl_department, id=department_id)

            job = Job(
                title=title,
                job_type=request.POST.get('job_type', '').strip(),
                description=request.POST.get('description', '').strip(),
                place=request.POST.get('place', '').strip(),
                salary=float(request.POST.get('salary', '0').strip()),
                responsibilities=request.POST.get('responsibilities', '').strip(),
                qualifications=request.POST.get('qualifications', '').strip(),
                vacancy=int(request.POST.get('vacancy', '0').strip()),
                passoutyear=int(request.POST.get('passoutyear', '').strip()),
                logo=request.FILES.get('logo'),
                department=department,
            )
            job.save()

            
            tutors = tbl_tutor.objects.filter(department=department)
            for tutor in tutors:
                notification_message = f"A new job '{job.title}' has been added in your department."
                TutorNotification.objects.create(
                    tutor=tutor,
                    job=job,
                    message=notification_message,
                )

            return JsonResponse({'success': True, 'message': 'Job added and notifications sent'})

        except Exception as e:
            print("Error:", e)
            return JsonResponse({'success': False, 'error': str(e)})
    else:
        departments = tbl_department.objects.all()
        return render(request, 'admin/admin_addjoblist.html', {'departments': departments})

from django.shortcuts import render, redirect
from .models import TutorNotification

def tutor_notifications(request):
    tutor_id = request.session.get('id')
    if not tutor_id:
        return redirect('tutor_login')  
    
    try:
        tutor = tbl_tutor.objects.get(id=tutor_id)
    except tbl_tutor.DoesNotExist:
        messages.error(request, "Tutor not found")
        return redirect('tutor_login')
    notifications = TutorNotification.objects.filter(tutor=tutor).order_by('-created_at')
    notifications.update(is_read=True)
    
    context = {
        'notifications': notifications,
    }
    return render(request, 'tutor/notifications.html', context)

def mark_notification_as_read(request, notification_id):
    try:
        notification = TutorNotification.objects.get(id=notification_id, tutor_id=request.session.get('tutor_id'))
        notification.is_read = True
        notification.save()
        return JsonResponse({'success': True})
    except TutorNotification.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Notification not found'})

def notifications_count(request):
    if request.user.is_authenticated:
        try:
            tutor = request.user.tutor 
            count = TutorNotification.objects.filter(tutor=tutor, is_read=False).count()
            return {'notifications_count': count}
        except Exception as e:
            print("Error fetching tutor notifications:", e)
            return {'notifications_count': 0}
    return {'notifications_count': 0}
